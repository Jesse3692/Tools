import openpyxl

def write_to_xlsx(path, sheet_name, value):
    index = len(value)
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    for i in range(0, index):
        for j in range(0, len(value[i])):
            sheet.cell(row=i+1, column=j+1, value=str(value[i][j]))
    workbook.save(path)
    print("xlsx表格成功写入数据")

def read_to_xlsx(path, sheet_name):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheet_name]
    for row in sheet.rows:
        for cell in row:
            print(cell.value, "\t", end="")
        print()

if __name__ == "__main__":
    book_name_xlsx = './sourcecode/file/test_excel.xlsx'
    sheet_name_xlsx = 'test_sheet'
    value_list = [
        ['姓名', '性别', '年龄', '城市'],
        ['张三', '男', '18', '深圳'],
        ['李四', '女', '20', '北京']
    ]
    # 写入数据
    write_to_xlsx(book_name_xlsx, sheet_name_xlsx, value_list)
    # 读数据
    read_to_xlsx(book_name_xlsx, sheet_name_xlsx)